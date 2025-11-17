#!/usr/bin/env python3
"""
Generate Bloomberg Excel workbook with BDH formulas for bulk data download.

Usage:
    python scripts/generate_bloomberg_workbook.py [--expanded] [--output PATH]

Options:
    --expanded          Use expanded 72-asset symbol map (default: 50-asset)
    --output PATH       Output path for Excel file (default: data/bloomberg/raw/bloomberg_export.xlsx)
    --start DATE        Start date (default: 1990-01-01)
    --end DATE          End date (default: 2025-11-16)

This script:
1. Reads symbol metadata from data/bloomberg/symbol_map.csv (or symbol_map_expanded.csv)
2. Creates an Excel workbook with BDH formula for bulk Bloomberg data download
3. Saves the workbook ready for use at Bloomberg Terminal

The generated Excel file contains:
- Column A: Pinnacle IDs (for reference)
- Column B: Asset Class (CM, EQ, FI, FX, etc.)
- Column C: Description
- Column D: Bloomberg Tickers
- Column E+: BDH formula that downloads all prices at once
"""

import argparse
import csv
from datetime import datetime
from pathlib import Path
from typing import Dict, List

import openpyxl


IGNORED_TICKER_TOKENS = {"<TBD>", "TBD"}


def _clean_field(value: str | None) -> str:
    """Trim whitespace and normalize None → empty string."""
    return (value or "").strip()


def load_symbol_map(csv_path: Path) -> List[Dict[str, str]]:
    """Load symbol map CSV and return list of symbol records."""
    symbols: List[Dict[str, str]] = []
    if not csv_path.exists():
        raise FileNotFoundError(f"Symbol map not found: {csv_path}")

    with csv_path.open(newline="") as fh:
        reader = csv.DictReader(fh)
        for raw in reader:
            pinnacle = _clean_field(raw.get("pinnacle_id"))
            ticker = _clean_field(raw.get("bloomberg_ticker"))

            if not pinnacle or not ticker:
                continue
            if pinnacle.startswith("#") or ticker.startswith("#"):
                continue
            if any(token in ticker.upper() for token in IGNORED_TICKER_TOKENS):
                continue

            record = {
                "pinnacle_id": pinnacle,
                "asset_class": _clean_field(raw.get("asset_class")),
                "description": _clean_field(raw.get("description")),
                "bloomberg_ticker": ticker,
                "notes": _clean_field(raw.get("notes")),
            }
            symbols.append(record)
    return symbols


def generate_workbook(
    symbols: list[dict],
    output_path: Path,
    start_date: str = "1990-01-01",
    end_date: str = "2025-11-16",
) -> None:
    """
    Generate Bloomberg Excel workbook with BDH formula.

    Args:
        symbols: List of symbol dictionaries from CSV
        output_path: Path to save Excel file
        start_date: Start date for BDH formula (YYYY-MM-DD)
        end_date: End date for BDH formula (YYYY-MM-DD)
    """
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bloomberg Data Export"

    # Write header row
    ws["A1"] = "Pinnacle ID"
    ws["B1"] = "Asset Class"
    ws["C1"] = "Description"
    ws["D1"] = "Bloomberg Ticker"
    ws["E1"] = "Notes"

    # Write symbol data (starting at row 2)
    for idx, symbol in enumerate(symbols, start=2):
        ws[f"A{idx}"] = symbol["pinnacle_id"]
        ws[f"B{idx}"] = symbol.get("asset_class", "")
        ws[f"C{idx}"] = symbol.get("description", "")
        ws[f"D{idx}"] = symbol["bloomberg_ticker"]
        ws[f"E{idx}"] = symbol.get("notes", "")

    # Format dates for Bloomberg (YYYYMMDD)
    start_bbg = start_date.replace("-", "")
    end_bbg = end_date.replace("-", "")

    # Calculate row range for BDH formula (row 2 to row 2+len(symbols)-1)
    first_row = 2
    last_row = first_row + len(symbols) - 1

    # Create BDH formula in cell F1 that downloads all symbols at once
    # Formula: =BDH($D$2:$D$N,"PX_LAST","YYYYMMDD","YYYYMMDD","Dir=V")
    bdh_formula = f'=BDH($D${first_row}:$D${last_row},"PX_LAST","{start_bbg}","{end_bbg}","Dir=V")'

    # Write the formula to F1
    ws["F1"] = bdh_formula

    # Add instruction text
    ws["A1"].comment = openpyxl.comments.Comment(
        text="Instructions:\n"
        "1. Open this file at Bloomberg Terminal\n"
        "2. Press Enter in cell F1 to execute BDH formula\n"
        "3. Wait 10-20 minutes for Bloomberg to load data\n"
        "4. Save and copy to USB drive\n"
        f"5. Run: python scripts/convert_bloomberg_to_parquet.py",
        author="X-Trend Data Export",
    )

    # Auto-size columns A-E for readability
    for col in ["A", "B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 25

    # Save workbook
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"✓ Generated Bloomberg workbook: {output_path}")
    print(f"  - {len(symbols)} symbols")
    print(f"  - Date range: {start_date} to {end_date}")
    print(f"  - BDH formula in cell F1")
    print()
    print("Next steps:")
    print("  1. Copy file to USB drive")
    print("  2. Open at Bloomberg Terminal")
    print("  3. Press Enter in cell F1 to load data")
    print("  4. Wait for Bloomberg to populate data (10-20 min)")
    print("  5. Save and bring back to convert to Parquet")


def main():
    parser = argparse.ArgumentParser(
        description="Generate Bloomberg Excel workbook for bulk data download"
    )
    parser.add_argument(
        "--expanded",
        action="store_true",
        help="Use expanded 72-asset symbol map (default: 50-asset)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output path for Excel file (default: auto-determined)",
    )
    parser.add_argument(
        "--start",
        type=str,
        default="1990-01-01",
        help="Start date (YYYY-MM-DD, default: 1990-01-01)",
    )
    parser.add_argument(
        "--end",
        type=str,
        default=datetime.now().strftime("%Y-%m-%d"),
        help="End date (YYYY-MM-DD, default: today)",
    )

    args = parser.parse_args()

    # Determine paths
    project_root = Path(__file__).parent.parent
    bloomberg_dir = project_root / "data" / "bloomberg"

    # Choose symbol map
    if args.expanded:
        symbol_map_path = bloomberg_dir / "symbol_map_expanded.csv"
        default_output = bloomberg_dir / "raw" / "bloomberg_export_72assets.xlsx"
    else:
        symbol_map_path = bloomberg_dir / "symbol_map.csv"
        default_output = bloomberg_dir / "raw" / "bloomberg_export.xlsx"

    output_path = args.output if args.output else default_output

    # Load symbols
    if not symbol_map_path.exists():
        print(f"ERROR: Symbol map not found: {symbol_map_path}")
        return 1

    symbols = load_symbol_map(symbol_map_path)

    if not symbols:
        print(f"ERROR: No symbols found in {symbol_map_path}")
        return 1

    print("=" * 70)
    print("Bloomberg Excel Workbook Generator")
    print("=" * 70)
    print()

    # Generate workbook
    generate_workbook(
        symbols=symbols,
        output_path=output_path,
        start_date=args.start,
        end_date=args.end,
    )

    return 0


if __name__ == "__main__":
    exit(main())
